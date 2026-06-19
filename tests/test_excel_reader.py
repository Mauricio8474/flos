import openpyxl
import pytest

from src.domain.models import ItemInventario
from src.infrastructure.adapters.excel_reader import ExcelFormulasAdapter, ExcelInventarioAdapter


class TestExcelInventarioAdapter:
    def test_archivo_inexistente(self):
        adapter = ExcelInventarioAdapter()
        with pytest.raises(FileNotFoundError):
            adapter.leer_inventario("no_existe.xlsx")

    def test_extension_invalida(self, tmp_path):
        csv_path = tmp_path / "datos.csv"
        csv_path.write_text("a,b,c")
        adapter = ExcelInventarioAdapter()
        with pytest.raises(ValueError, match="Formato no soportado"):
            adapter.leer_inventario(str(csv_path))

    def test_excel_vacio(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.save(tmp_path / "vacio.xlsx")
        adapter = ExcelInventarioAdapter()
        with pytest.raises(ValueError, match="no contiene datos"):
            adapter.leer_inventario(str(tmp_path / "vacio.xlsx"))

    def test_faltan_columnas(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["SKU", "Nombre"])
        wb.save(tmp_path / "sin_cantidad.xlsx")
        adapter = ExcelInventarioAdapter()
        with pytest.raises(ValueError, match="no encontrada"):
            adapter.leer_inventario(str(tmp_path / "sin_cantidad.xlsx"))

    def test_lectura_correcta(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["SKU", "Nombre", "Cantidad_KG"])
        ws.append(["MP001", "Material 1", 100.0])
        ws.append(["MP002", "Material 2", 50.5])
        wb.save(tmp_path / "inventario.xlsx")
        adapter = ExcelInventarioAdapter()
        resultado = adapter.leer_inventario(str(tmp_path / "inventario.xlsx"))
        assert len(resultado) == 2
        assert all(isinstance(i, ItemInventario) for i in resultado)
        assert resultado[0].sku == "MP001"
        assert resultado[0].cantidad_kg == 100.0
        assert resultado[1].sku == "MP002"

    def test_sku_duplicado_se_acumula(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["SKU", "Nombre", "Cantidad_KG"])
        ws.append(["MP001", "Material 1", 50.0])
        ws.append(["MP001", "Material 1 extra", 30.0])
        wb.save(tmp_path / "duplicados.xlsx")
        adapter = ExcelInventarioAdapter()
        resultado = adapter.leer_inventario(str(tmp_path / "duplicados.xlsx"))
        items = {i.sku: i for i in resultado}
        assert items["MP001"].cantidad_kg == 80.0

    def test_formato_siigo(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        for _ in range(6):
            ws.append([])
        ws.append(["", "", "", "", "", "", "REFERENCIA", "DESCRIPCION", "", "SALDO", "", "COSTO UNITARIO"])
        ws.append(["", "", "", "", "", "", "MP001", "Material A", "", "100.5", "", "25.00"])
        wb.save(tmp_path / "siigo.xlsx")
        adapter = ExcelInventarioAdapter(
            fila_encabezados=7,
            mapeo={"SKU": "REFERENCIA", "Nombre": "DESCRIPCION", "Cantidad_KG": "SALDO", "CostoUnitario": "COSTO UNITARIO"},
        )
        resultado = adapter.leer_inventario(str(tmp_path / "siigo.xlsx"))
        assert len(resultado) == 1
        assert resultado[0].sku == "MP001"
        assert resultado[0].cantidad_kg == 100.5
        assert resultado[0].costo_unitario == 25.0

    def test_cantidad_invalida_se_trata_como_cero(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["SKU", "Nombre", "Cantidad_KG"])
        ws.append(["MP001", "Material 1", "invalido"])
        wb.save(tmp_path / "invalido.xlsx")
        adapter = ExcelInventarioAdapter()
        resultado = adapter.leer_inventario(str(tmp_path / "invalido.xlsx"))
        assert resultado[0].cantidad_kg == 0.0

    def test_celda_vacia_en_cantidad(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["SKU", "Nombre", "Cantidad_KG"])
        ws.append(["MP001", "Material 1", None])
        wb.save(tmp_path / "vacio_cantidad.xlsx")
        adapter = ExcelInventarioAdapter()
        resultado = adapter.leer_inventario(str(tmp_path / "vacio_cantidad.xlsx"))
        assert resultado[0].cantidad_kg == 0.0


class TestExcelFormulasAdapter:
    def test_archivo_inexistente(self):
        adapter = ExcelFormulasAdapter()
        with pytest.raises(FileNotFoundError):
            adapter.leer_formulas("no_existe.xlsx")

    def test_lectura_correcta(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "formulas"
        ws.append(["col0", "Ref Prod Terminado", "col2", "MP POR REF", "MP (nombre)", "col5"])
        ws.append(["", "F001", "", "SKU001", "Producto A", ""])
        ws.append(["", "F001", "", "SKU002", "Producto A", ""])
        ws.append(["", "F002", "", "SKU003", "Producto B", ""])
        wb.save(tmp_path / "formulas.xlsx")
        adapter = ExcelFormulasAdapter(mapeo={"ID": 1, "MP": 4, "SKU": 3, "KG": 11})
        with pytest.raises(ValueError, match="Estructura"):
            adapter.leer_formulas(str(tmp_path / "formulas.xlsx"))

    def test_lectura_correcta_con_kg(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "formulas"
        row = ["", "", "", "", "", "", "", "", "", "", "", "KG"]
        for _ in range(11):
            row.append("")
        row[11] = "KG"
        ws.append(row)
        ws.append(["", "F001", "", "SKU001", "Producto A", "", "", "", "", "", "", 50.0])
        wb.save(tmp_path / "formulas_con_kg.xlsx")
        adapter = ExcelFormulasAdapter(mapeo={"ID": 1, "MP": 4, "SKU": 3, "KG": 11})
        resultado = adapter.leer_formulas(str(tmp_path / "formulas_con_kg.xlsx"))
        assert "F001" in resultado
        f = resultado["F001"]
        assert f.nombre == "Producto A"
        assert len(f.componentes) == 1
        assert f.componentes[0].sku == "SKU001"
        assert f.componentes[0].porcentaje == 50.0

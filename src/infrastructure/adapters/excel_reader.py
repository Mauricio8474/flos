from pathlib import Path

import openpyxl

from src.application.use_cases import PuertoFormulas, PuertoInventario
from src.domain.models import ComponenteFormula, Formula, ItemInventario


class ExcelInventarioAdapter(PuertoInventario):
    def __init__(
        self,
        fila_encabezados: int = 1,
        mapeo: dict[str, str] | None = None,
    ) -> None:
        self._fila_encabezados = fila_encabezados
        self._mapeo = mapeo or {}

    def leer_inventario(self, ruta: str) -> list[ItemInventario]:
        archivo = Path(ruta)

        if not archivo.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta}")
        if archivo.suffix not in (".xlsx", ".xls"):
            raise ValueError(f"Formato no soportado: {archivo.suffix}")

        libro = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
        hoja = libro.active

        filas = list(hoja.iter_rows(values_only=True))
        if not filas:
            raise ValueError("El archivo Excel no contiene datos")

        idx_header = self._fila_encabezados - 1
        if idx_header >= len(filas):
            raise ValueError(
                f"Fila de encabezados {self._fila_encabezados} no existe "
                f"(el archivo tiene {len(filas)} filas)"
            )

        encabezados = [str(c).strip() if c is not None else "" for c in filas[idx_header]]

        col_sku = self._mapeo.get("SKU", "SKU")
        col_nombre = self._mapeo.get("Nombre", "Nombre")
        col_cant = self._mapeo.get("Cantidad_KG", "Cantidad_KG")
        col_costo = self._mapeo.get("CostoUnitario", "CostoUnitario")

        if col_sku not in encabezados:
            raise ValueError(f"Columna SKU ('{col_sku}') no encontrada. Columnas: {encabezados}")
        if col_cant not in encabezados:
            raise ValueError(f"Columna de cantidad ('{col_cant}') no encontrada. Columnas: {encabezados}")

        idx_sku = encabezados.index(col_sku)
        idx_nombre = encabezados.index(col_nombre) if col_nombre in encabezados else None
        idx_cant = encabezados.index(col_cant)
        idx_costo = encabezados.index(col_costo) if col_costo in encabezados else None

        agrupados: dict[str, ItemInventario] = {}

        for fila in filas[idx_header + 1:]:
            sku = str(fila[idx_sku]).strip() if fila[idx_sku] is not None else ""
            if not sku:
                continue

            nombre = str(fila[idx_nombre]).strip() if idx_nombre is not None and fila[idx_nombre] is not None else ""

            try:
                cantidad = float(fila[idx_cant]) if fila[idx_cant] is not None else 0.0
            except (ValueError, TypeError):
                cantidad = 0.0

            try:
                costo = float(fila[idx_costo]) if idx_costo is not None and fila[idx_costo] is not None else 0.0
            except (ValueError, TypeError):
                costo = 0.0

            if sku in agrupados:
                existente = agrupados[sku]
                agrupados[sku] = ItemInventario(
                    sku=sku,
                    nombre=nombre or existente.nombre,
                    cantidad_kg=round(existente.cantidad_kg + cantidad, 3),
                    costo_unitario=costo or existente.costo_unitario,
                )
            else:
                agrupados[sku] = ItemInventario(
                    sku=sku, nombre=nombre, cantidad_kg=round(cantidad, 3), costo_unitario=round(costo, 2),
                )

        return list(agrupados.values())


class ExcelFormulasAdapter(PuertoFormulas):
    HOJA = "formulas"
    COL_ID = 1   # B  — REFERENCIA PRODUCTO TERMINADO
    COL_MP = 4   # E  — MP (nombre)
    COL_SKU = 3  # D  — MP POR REF
    COL_KG = 11  # L  — fórmula en Kg

    def __init__(self, mapeo: dict[str, int] | None = None) -> None:
        self._mapeo = mapeo or {}

    def leer_formulas(self, ruta: str) -> dict[str, Formula]:
        archivo = Path(ruta)

        if not archivo.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta}")

        libro = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
        hoja = libro[self.HOJA]

        filas = list(hoja.iter_rows(values_only=True))
        if not filas:
            raise ValueError("El archivo Excel no contiene datos")

        col_id = self._mapeo.get("ID", self.COL_ID)
        col_mp = self._mapeo.get("MP", self.COL_MP)
        col_sku = self._mapeo.get("SKU", self.COL_SKU)
        col_kg = self._mapeo.get("KG", self.COL_KG)

        max_col = max(col_id, col_mp, col_sku, col_kg)
        encabezados = [str(c).strip() if c is not None else "" for c in filas[0]]
        if len(encabezados) <= max_col:
            raise ValueError(
                f"Estructura de columnas inesperada. "
                f"Se esperaban al menos {max_col + 1} columnas, se encontraron {len(encabezados)}"
            )

        agrupadas: dict[str, list[ComponenteFormula]] = {}
        nombres: dict[str, str] = {}

        for fila in filas[1:]:
            ref = str(fila[col_id]).strip() if fila[col_id] is not None else ""
            if not ref:
                continue

            if ref not in nombres:
                mp = str(fila[col_mp]).strip() if fila[col_mp] is not None else ref
                nombres[ref] = mp

            sku = str(fila[col_sku]).strip() if fila[col_sku] is not None else ""
            if not sku:
                continue

            kg = fila[col_kg]
            try:
                porcentaje = float(kg) if kg is not None else 0.0
            except (ValueError, TypeError):
                porcentaje = 0.0

            if ref not in agrupadas:
                agrupadas[ref] = []
            agrupadas[ref].append(ComponenteFormula(sku=sku, porcentaje=porcentaje))

        return {
            ref: Formula(nombre=nombres[ref], componentes=tuple(comps))
            for ref, comps in agrupadas.items()
        }

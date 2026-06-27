import io
import logging
import os
from pathlib import Path
from tempfile import NamedTemporaryFile

import openpyxl
import uvicorn
from fastapi import Depends, FastAPI, File, Form, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse
from pydantic import BaseModel
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph

from src.infrastructure.auth import (
    crear_token,
    hash_password,
    obtener_usuario_actual,
    requerir_rol,
    verificar_password,
)

from src.domain.models import ComponenteFormula, Formula, ItemInventario
from src.domain.services import CalculadorMRP
from src.application.use_cases import (
    CalcularExplosion, CalcularExplosionBatch, CambiarEstadoOrden,
    CrearControlCalidad, CrearLote, GenerarAlertasStock,
    GenerarSugerenciasCompra, RegistrarResultadoControl, TrazarLote,
)
from src.infrastructure.adapters.excel_reader import ExcelFormulasAdapter, ExcelInventarioAdapter
from src.infrastructure.adapters.repositories import (
    PostgresRepositorioAuditoria,
    PostgresRepositorioControlCalidad,
    PostgresRepositorioFormula,
    PostgresRepositorioInventario,
    PostgresRepositorioLotes,
    PostgresRepositorioOrdenes,
    PostgresRepositorioUsuario,
    init_db,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger("flos")

app = FastAPI(title="Flos MES", version="0.1.0")

app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])


STATIC_DIR = Path(__file__).parent / "static"


@app.get("/")
def root():
    return RedirectResponse(url="/dashboard/")


# Dashboard catch-all — keep this AFTER any /dashboard/* specific routes
@app.get("/dashboard/estadisticas", dependencies=[Depends(obtener_usuario_actual)])
def obtener_estadisticas() -> dict:
    return _repo_ordenes().estadisticas()


@app.get("/dashboard/{rest:path}")
def dashboard():
    return FileResponse(str(STATIC_DIR / "index.html"))


# ---------------------------------------------------------------------------
# DTOs
# ---------------------------------------------------------------------------
class ComponenteInput(BaseModel):
    sku: str
    porcentaje: float


class FormulaInput(BaseModel):
    id: str
    nombre: str
    componentes: list[ComponenteInput]


class LoginInput(BaseModel):
    username: str
    password: str


class UsuarioInput(BaseModel):
    username: str
    password: str
    rol: str
    nombre: str


class OrdenBatch(BaseModel):
    id_formula: str
    cantidad: float


# ---------------------------------------------------------------------------
# Startup — repos + seed admin
# ---------------------------------------------------------------------------
@app.on_event("startup")
def startup():
    import os
    jwt_secret = os.getenv("JWT_SECRET")
    if not jwt_secret or jwt_secret in ("change-this-secret-in-production", "cambiar-esta-clave-en-produccion"):
        logger.warning("JWT_SECRET no configurado o usa valor por defecto. ¡Cambiar en producción!")
    db_url = os.getenv("DATABASE_URL")
    if not db_url:
        logger.error("DATABASE_URL no configurada")
        raise RuntimeError("DATABASE_URL environment variable is required")
    sf = init_db(db_url)
    app.state.sf = sf
    app.state.repo_formula = PostgresRepositorioFormula(sf)
    app.state.repo_inventario = PostgresRepositorioInventario(sf)
    app.state.repo_auditoria = PostgresRepositorioAuditoria(sf)
    app.state.repo_usuarios = PostgresRepositorioUsuario(sf)
    app.state.repo_ordenes = PostgresRepositorioOrdenes(sf)
    app.state.repo_control_calidad = PostgresRepositorioControlCalidad(sf)
    app.state.repo_lotes = PostgresRepositorioLotes(sf)
    app.state.caso_calcular_explosion = CalcularExplosion(
        app.state.repo_formula, app.state.repo_inventario, app.state.repo_ordenes,
    )
    app.state.caso_calcular_explosion_batch = CalcularExplosionBatch(
        app.state.repo_formula, app.state.repo_inventario, app.state.repo_ordenes,
    )
    app.state.caso_cambiar_estado_orden = CambiarEstadoOrden(app.state.repo_ordenes)
    app.state.caso_generar_alertas = GenerarAlertasStock(app.state.repo_inventario)
    app.state.caso_generar_sugerencias = GenerarSugerenciasCompra(app.state.repo_inventario, app.state.repo_ordenes)
    app.state.caso_crear_control = CrearControlCalidad(app.state.repo_control_calidad)
    app.state.caso_registrar_resultado = RegistrarResultadoControl(app.state.repo_control_calidad)
    app.state.caso_crear_lote = CrearLote(app.state.repo_lotes, app.state.repo_ordenes)
    app.state.caso_trazar_lote = TrazarLote(app.state.repo_lotes, app.state.repo_ordenes)

    repo_usu = app.state.repo_usuarios
    if not repo_usu.existe_admin():
        repo_usu.guardar("admin", hash_password("123456789"), "admin", "Administrador")
        logger.info("Usuario admin creado con contraseña por defecto")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _repo_formula():
    return app.state.repo_formula


def _repo_inventario():
    return app.state.repo_inventario


def _repo_auditoria():
    return app.state.repo_auditoria


def _repo_usuarios():
    return app.state.repo_usuarios


def _repo_ordenes():
    return app.state.repo_ordenes


def _paginar(items: list | dict, total: int, page: int, page_size: int) -> dict:
    if page_size <= 0:
        page_size = total if total > 0 else 1
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": (total + page_size - 1) // page_size if page_size > 0 else 1,
    }


def _caso_calcular_explosion():
    return app.state.caso_calcular_explosion


def _caso_calcular_explosion_batch():
    return app.state.caso_calcular_explosion_batch


def _caso_cambiar_estado_orden():
    return app.state.caso_cambiar_estado_orden


def _caso_generar_alertas():
    return app.state.caso_generar_alertas


def _caso_generar_sugerencias():
    return app.state.caso_generar_sugerencias


def _repo_control():
    return app.state.repo_control_calidad


def _repo_lotes():
    return app.state.repo_lotes


def _caso_crear_control():
    return app.state.caso_crear_control


def _caso_registrar_resultado():
    return app.state.caso_registrar_resultado


def _caso_crear_lote():
    return app.state.caso_crear_lote


def _caso_trazar_lote():
    return app.state.caso_trazar_lote


def _auditar(entidad: str, entidad_id: str, accion: str, detalle: str, usuario: str) -> None:
    _repo_auditoria().registrar(entidad, entidad_id, accion, detalle, usuario)


# ---------------------------------------------------------------------------
# Autenticación
# ---------------------------------------------------------------------------
@app.post("/auth/login")
def login(body: LoginInput) -> dict:
    user = _repo_usuarios().obtener(body.username)
    if not user or not user.get("activo"):
        return {"error": "Usuario o contraseña incorrectos"}
    if not verificar_password(body.password, user["password_hash"]):
        return {"error": "Usuario o contraseña incorrectos"}
    token = crear_token(user["username"], user["rol"], user["nombre"])
    return {"access_token": token, "token_type": "bearer", "rol": user["rol"], "nombre": user["nombre"]}


@app.get("/auth/me")
def auth_me(usuario: dict = Depends(obtener_usuario_actual)) -> dict:
    return {"username": usuario["sub"], "rol": usuario["rol"], "nombre": usuario["nombre"]}


# ---------------------------------------------------------------------------
# Usuarios (solo admin)
# ---------------------------------------------------------------------------
@app.get("/usuarios", dependencies=[Depends(requerir_rol("admin"))])
def listar_usuarios(q: str = "") -> list[dict]:
    return _repo_usuarios().listar(q)


@app.post("/usuarios", dependencies=[Depends(requerir_rol("admin"))])
def crear_usuario(body: UsuarioInput) -> dict:
    if body.rol not in ("admin", "ingenieria", "almacen", "produccion", "consultor"):
        return {"error": f"Rol invalido: {body.rol}"}
    existe = _repo_usuarios().obtener(body.username)
    if existe:
        return {"error": f"Usuario '{body.username}' ya existe"}
    _repo_usuarios().guardar(body.username, hash_password(body.password), body.rol, body.nombre)
    return {"mensaje": f"Usuario '{body.username}' creado", "rol": body.rol}


@app.put("/usuarios/{username}", dependencies=[Depends(requerir_rol("admin"))])
def actualizar_usuario(username: str, body: UsuarioInput) -> dict:
    existe = _repo_usuarios().obtener(username)
    if not existe:
        return {"error": f"Usuario '{username}' no encontrado"}
    _repo_usuarios().guardar(username, hash_password(body.password), body.rol, body.nombre)
    return {"mensaje": f"Usuario '{username}' actualizado"}


@app.delete("/usuarios/{username}", dependencies=[Depends(requerir_rol("admin"))])
def eliminar_usuario(username: str) -> dict:
    repo = _repo_usuarios()
    user = repo.obtener(username)
    if not user:
        return {"error": f"Usuario '{username}' no encontrado"}
    repo.guardar(username, user["password_hash"], user["rol"], user["nombre"], activo=False)
    return {"mensaje": f"Usuario '{username}' desactivado"}


# ---------------------------------------------------------------------------
# Fórmulas
# ---------------------------------------------------------------------------
@app.get("/produccion/formulas", dependencies=[Depends(obtener_usuario_actual)])
def listar_formulas(page: int = 0, page_size: int = 0, q: str = "") -> dict:
    todas, total = _repo_formula().listar(page, page_size, q)
    items = [{"id": ref, "nombre": f.nombre, "componentes": [{"sku": c.sku, "porcentaje": c.porcentaje} for c in f.componentes]} for ref, f in sorted(todas.items())]
    if not page:
        page = 1
        page_size = total if total > 0 else 1
    return _paginar(items, total, page, page_size)


@app.get("/produccion/formulas/excel", dependencies=[Depends(obtener_usuario_actual)])
def descargar_formulas_excel() -> StreamingResponse:
    todas, _ = _repo_formula().listar()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "formulas"
    ws.append(["REFERENCIA PRODUCTO TERMINADO", "MP", "MP POR REF", "fórmula en Kg"])
    for ref, f in sorted(todas.items()):
        for c in f.componentes:
            ws.append([ref, f.nombre, c.sku, c.porcentaje])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=formulas.xlsx"},
    )


@app.get("/produccion/formulas/{id_formula}", dependencies=[Depends(obtener_usuario_actual)])
def obtener_formula(id_formula: str) -> dict:
    f = _repo_formula().obtener(id_formula)
    if not f:
        return {"error": f"Formula '{id_formula}' no encontrada"}
    return {"id": id_formula, "nombre": f.nombre, "componentes": [{"sku": c.sku, "porcentaje": c.porcentaje} for c in f.componentes]}


@app.post("/produccion/formulas", dependencies=[Depends(requerir_rol("admin", "ingenieria"))])
def crear_formula(body: FormulaInput, usuario: dict = Depends(obtener_usuario_actual)) -> dict:
    repo = _repo_formula()
    if repo.obtener(body.id):
        return {"error": f"La formula '{body.id}' ya existe"}
    repo.guardar(body.id, Formula(nombre=body.nombre, componentes=tuple(ComponenteFormula(sku=c.sku, porcentaje=c.porcentaje) for c in body.componentes)))
    _auditar("formula", body.id, "CREAR", f"{len(body.componentes)} componentes", usuario["sub"])
    return {"mensaje": f"Formula '{body.id}' creada", "componentes": len(body.componentes)}


@app.put("/produccion/formulas/{id_formula}", dependencies=[Depends(requerir_rol("admin", "ingenieria"))])
def actualizar_formula(id_formula: str, body: FormulaInput, usuario: dict = Depends(obtener_usuario_actual)) -> dict:
    repo = _repo_formula()
    if not repo.obtener(id_formula):
        return {"error": f"Formula '{id_formula}' no encontrada"}
    repo.guardar(id_formula, Formula(nombre=body.nombre, componentes=tuple(ComponenteFormula(sku=c.sku, porcentaje=c.porcentaje) for c in body.componentes)))
    _auditar("formula", id_formula, "ACTUALIZAR", f"{len(body.componentes)} componentes", usuario["sub"])
    return {"mensaje": f"Formula '{id_formula}' actualizada", "componentes": len(body.componentes)}


@app.delete("/produccion/formulas/{id_formula}", dependencies=[Depends(requerir_rol("admin", "ingenieria"))])
def eliminar_formula(id_formula: str, usuario: dict = Depends(obtener_usuario_actual)) -> dict:
    if _repo_formula().eliminar(id_formula):
        _auditar("formula", id_formula, "ELIMINAR", "", usuario["sub"])
        return {"mensaje": f"Formula '{id_formula}' eliminada"}
    return {"error": f"Formula '{id_formula}' no encontrada"}


@app.post("/produccion/cargar-formulas", dependencies=[Depends(requerir_rol("admin", "ingenieria"))])
def cargar_formulas(
    archivo: UploadFile = File(...),
    columna_id: int | None = Form(None),
    columna_mp: int | None = Form(None),
    columna_sku: int | None = Form(None),
    columna_kg: int | None = Form(None),
    usuario: dict = Depends(obtener_usuario_actual),
) -> dict:
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(archivo.file.read())
        tmp_path = tmp.name

    mapeo = {}
    if columna_id is not None:
        mapeo["ID"] = columna_id
    if columna_mp is not None:
        mapeo["MP"] = columna_mp
    if columna_sku is not None:
        mapeo["SKU"] = columna_sku
    if columna_kg is not None:
        mapeo["KG"] = columna_kg

    adapter = ExcelFormulasAdapter(mapeo=mapeo)
    formulas = adapter.leer_formulas(tmp_path)
    _repo_formula().guardar_muchos(formulas)
    os.unlink(tmp_path)
    _auditar("formula", "MASIVO", "CARGAR", f"{len(formulas)} formulas desde Excel", usuario["sub"])
    return {"mensaje": f"{len(formulas)} formulas cargadas"}


# ---------------------------------------------------------------------------
# Inventario
# ---------------------------------------------------------------------------
@app.post("/produccion/cargar-inventario", dependencies=[Depends(requerir_rol("admin", "ingenieria", "almacen"))])
def cargar_inventario(
    archivo: UploadFile = File(...),
    fila_encabezados: int = Form(1),
    columna_sku: str | None = Form(None),
    columna_nombre: str | None = Form(None),
    columna_cantidad: str | None = Form(None),
    columna_costo: str | None = Form(None),
    usuario: dict = Depends(obtener_usuario_actual),
) -> dict:
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(archivo.file.read())
        tmp_path = tmp.name

    mapeo = {}
    if columna_sku:
        mapeo["SKU"] = columna_sku
    if columna_nombre:
        mapeo["Nombre"] = columna_nombre
    if columna_cantidad:
        mapeo["Cantidad_KG"] = columna_cantidad
    if columna_costo:
        mapeo["CostoUnitario"] = columna_costo

    adapter = ExcelInventarioAdapter(fila_encabezados=fila_encabezados, mapeo=mapeo)
    inventario = adapter.leer_inventario(tmp_path)
    _repo_inventario().guardar_muchos(inventario)
    _auditar("inventario", "MASIVO", "CARGAR", f"{len(inventario)} SKUs desde Excel", usuario["sub"])
    return {"mensaje": f"Inventario actualizado: {len(inventario)} SKUs"}


@app.get("/produccion/inventario", dependencies=[Depends(obtener_usuario_actual)])
def obtener_inventario(page: int = 0, page_size: int = 0, q: str = "") -> dict:
    items, total = _repo_inventario().obtener_todos(page, page_size, q)
    data = [
        {
            "sku": i.sku,
            "nombre": i.nombre,
            "cantidad_kg": i.cantidad_kg,
            "costo_unitario": i.costo_unitario,
        }
        for i in items
    ]
    if not page:
        page = 1
        page_size = total if total > 0 else 1
    return _paginar(data, total, page, page_size)


@app.put("/produccion/inventario/{sku}", dependencies=[Depends(requerir_rol("admin", "ingenieria", "almacen"))])
def actualizar_stock_minimo(sku: str, stock_minimo: float = Form(...)) -> dict:
    item = _repo_inventario().obtener(sku)
    if not item:
        return {"error": f"SKU '{sku}' no encontrado"}
    _repo_inventario().actualizar_stock_minimo(sku, stock_minimo)
    return {"mensaje": f"Stock mínimo de '{sku}' actualizado a {stock_minimo} kg"}


# ---------------------------------------------------------------------------
# Alertas de stock
# ---------------------------------------------------------------------------
@app.get("/alertas/stock", dependencies=[Depends(obtener_usuario_actual)])
def listar_alertas_stock() -> list[dict]:
    return _caso_generar_alertas().ejecutar()


# ---------------------------------------------------------------------------
# Sugerencias de compra
# ---------------------------------------------------------------------------
@app.get("/sugerencias-compra", dependencies=[Depends(obtener_usuario_actual)])
def listar_sugerencias_compra() -> dict:
    return {"sugerencias": _caso_generar_sugerencias().ejecutar()}


# ---------------------------------------------------------------------------
# Auditoría
# ---------------------------------------------------------------------------
@app.get("/produccion/auditoria", dependencies=[Depends(requerir_rol("admin"))])
def listar_auditoria(page: int = 0, page_size: int = 0, q: str = "") -> dict:
    items, total = _repo_auditoria().listar(page, page_size, q)
    if not page:
        page = 1
        page_size = total if total > 0 else 1
    return _paginar(items, total, page, page_size)


def _inventario_a_dict(items: list) -> dict[str, ItemInventario]:
    return {i.sku: i for i in items}


# ---------------------------------------------------------------------------
# Explosión
# ---------------------------------------------------------------------------
@app.post("/produccion/calcular-explosion", dependencies=[Depends(requerir_rol("admin", "ingenieria", "produccion"))])
def calcular_explosion(
    id_formula: str = Form(...),
    cantidad_a_producir_kg: float = Form(...),
    usuario: dict = Depends(obtener_usuario_actual),
) -> list[dict]:
    resultados, error = _caso_calcular_explosion().ejecutar(
        id_formula, cantidad_a_producir_kg, usuario["sub"],
    )
    if error:
        return [{"error": error}]
    return resultados


@app.post("/produccion/calcular-explosion/batch", dependencies=[Depends(requerir_rol("admin", "ingenieria", "produccion"))])
def calcular_explosion_batch(
    ordenes: list[OrdenBatch],
    usuario: dict = Depends(obtener_usuario_actual),
) -> list[dict]:
    return _caso_calcular_explosion_batch().ejecutar(
        [{"id_formula": o.id_formula, "cantidad": o.cantidad} for o in ordenes],
        usuario["sub"],
    )


@app.post("/produccion/calcular-explosion/batch/excel", dependencies=[Depends(requerir_rol("admin", "ingenieria", "produccion"))])
def calcular_explosion_batch_excel(
    archivo: UploadFile = File(...),
    columna_id: int | None = Form(None),
    columna_cantidad: int | None = Form(None),
    usuario: dict = Depends(obtener_usuario_actual),
) -> list[dict]:
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(archivo.file.read())
        tmp_path = tmp.name
    wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    os.unlink(tmp_path)
    if not filas:
        return [{"error": "El archivo Excel no contiene datos"}]
    encabezados = [str(c).strip() if c is not None else "" for c in filas[0]]
    idx_id = columna_id if columna_id is not None else 0
    idx_cant = columna_cantidad if columna_cantidad is not None else 1
    if idx_id >= len(encabezados) or idx_cant >= len(encabezados):
        return [{"error": f"Índices de columna fuera de rango. El archivo tiene {len(encabezados)} columnas"}]
    ordenes = []
    for fila in filas[1:]:
        ref = str(fila[idx_id]).strip() if fila[idx_id] is not None else ""
        try:
            cant = float(fila[idx_cant]) if fila[idx_cant] is not None else 0.0
        except (ValueError, TypeError):
            cant = 0.0
        if ref and cant > 0:
            ordenes.append({"id_formula": ref, "cantidad": cant})
    if not ordenes:
        return [{"error": "No se encontraron órdenes válidas en el Excel"}]
    return _caso_calcular_explosion_batch().ejecutar(ordenes, usuario["sub"])


def _generar_excel_resultados(resultados, id_formula) -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Explosion"
    ws.append(["SKU", "Nombre", "Requerido_KG", "Disponible_KG", "Faltante_KG", "Cubierto", "Nota"])
    for r in resultados:
        ws.append([r.sku, r.nombre, r.requerido_kg, r.disponible_kg, r.faltante_kg, r.cubierto, r.nota])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=explosion_{id_formula}.xlsx"})


def _generar_pdf_resultados(resultados, id_formula, cantidad) -> StreamingResponse:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph(f"Explosion de Materiales: {id_formula}", estilos["Title"]),
        Paragraph(f"Cantidad a producir: {cantidad} kg", estilos["Normal"]),
        Paragraph("<br/>", estilos["Normal"]),
    ]
    data = [["SKU", "Nombre", "Requerido (kg)", "Disponible (kg)", "Faltante (kg)", "Cubierto", "Nota"]]
    for r in resultados:
        data.append([r.sku, r.nombre, str(r.requerido_kg), str(r.disponible_kg), str(r.faltante_kg), "Si" if r.cubierto else "No", r.nota])
    tabla = Table(data)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#DCE6F1")]),
    ]))
    elementos.append(tabla)
    doc.build(elementos)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=explosion_{id_formula}.pdf"})


def _explosion_resultados(formula, cantidad, inventario):
    return CalculadorMRP.calcular_explosion(formula, cantidad, inventario), None


@app.post("/produccion/calcular-explosion/excel", dependencies=[Depends(requerir_rol("admin", "ingenieria", "almacen", "produccion"))])
def calcular_explosion_excel(id_formula: str = Form(...), cantidad_a_producir_kg: float = Form(...)) -> StreamingResponse:
    formula = _repo_formula().obtener(id_formula)
    if not formula:
        return StreamingResponse(iter([b"Formula no encontrada"]), media_type="text/plain", status_code=404)
    resultados, _ = _explosion_resultados(formula, cantidad_a_producir_kg, _inventario_a_dict(_repo_inventario().obtener_todos()[0]))
    return _generar_excel_resultados(resultados, id_formula)


@app.post("/produccion/calcular-explosion/pdf", dependencies=[Depends(requerir_rol("admin", "ingenieria", "almacen", "produccion"))])
def calcular_explosion_pdf(id_formula: str = Form(...), cantidad_a_producir_kg: float = Form(...)) -> StreamingResponse:
    formula = _repo_formula().obtener(id_formula)
    if not formula:
        return StreamingResponse(iter([b"Formula no encontrada"]), media_type="text/plain", status_code=404)
    resultados, _ = _explosion_resultados(formula, cantidad_a_producir_kg, _inventario_a_dict(_repo_inventario().obtener_todos()[0]))
    return _generar_pdf_resultados(resultados, id_formula, cantidad_a_producir_kg)


# ---------------------------------------------------------------------------
# Órdenes de producción
# ---------------------------------------------------------------------------
@app.get("/ordenes", dependencies=[Depends(obtener_usuario_actual)])
def listar_ordenes(page: int = 0, page_size: int = 0, q: str = "") -> dict:
    items, total = _repo_ordenes().listar(page, page_size, q)
    if not page:
        page = 1
        page_size = total if total > 0 else 1
    return _paginar(items, total, page, page_size)


@app.get("/ordenes/{id_orden}", dependencies=[Depends(obtener_usuario_actual)])
def obtener_orden(id_orden: str) -> dict:
    orden = _repo_ordenes().obtener(id_orden)
    if not orden:
        return {"error": f"Orden '{id_orden}' no encontrada"}
    return orden


@app.delete("/ordenes/{id_orden}", dependencies=[Depends(requerir_rol("admin"))])
def eliminar_orden(id_orden: str) -> dict:
    orden = _repo_ordenes().obtener(id_orden)
    if not orden:
        return {"error": f"Orden '{id_orden}' no encontrada"}
    if orden.get("estado") != "pendiente":
        return {"error": "Solo se pueden eliminar órdenes en estado 'pendiente'"}
    if _repo_ordenes().eliminar(id_orden):
        return {"mensaje": f"Orden '{id_orden}' eliminada"}
    return {"error": f"Orden '{id_orden}' no encontrada"}


@app.post("/ordenes/{id_orden}/iniciar", dependencies=[Depends(requerir_rol("admin", "produccion"))])
def iniciar_orden(id_orden: str, usuario: dict = Depends(obtener_usuario_actual)) -> dict:
    res = _caso_cambiar_estado_orden().ejecutar(id_orden, "en_produccion")
    if "error" in res:
        return res
    _auditar("orden", id_orden, "INICIAR", "Orden pasa a en_produccion", usuario["sub"])
    return res


@app.post("/ordenes/{id_orden}/completar", dependencies=[Depends(requerir_rol("admin", "produccion"))])
def completar_orden(id_orden: str, usuario: dict = Depends(obtener_usuario_actual)) -> dict:
    res = _caso_cambiar_estado_orden().ejecutar(id_orden, "completada")
    if "error" in res:
        return res
    _auditar("orden", id_orden, "COMPLETAR", "Orden completada - inventario consumido", usuario["sub"])
    return res


# ---------------------------------------------------------------------------
# Control de calidad
# ---------------------------------------------------------------------------
@app.get("/ordenes/{id_orden}/controles", dependencies=[Depends(obtener_usuario_actual)])
def listar_controles_orden(id_orden: str) -> list[dict]:
    return _repo_control().listar_por_orden(id_orden)


@app.post("/ordenes/{id_orden}/controles", dependencies=[Depends(requerir_rol("admin", "produccion"))])
def crear_control_calidad(id_orden: str, tipo: str = Form(...), usuario: dict = Depends(obtener_usuario_actual)) -> dict:
    res = _caso_crear_control().ejecutar(id_orden, tipo)
    if "error" in res:
        return res
    _auditar("control_calidad", res["id"], "CREAR", f"Control '{tipo}' para orden {id_orden[:8]}…", usuario["sub"])
    return res


@app.put("/control-calidad/{id_control}", dependencies=[Depends(requerir_rol("admin", "produccion"))])
def actualizar_resultado_control(id_control: str, resultado: str = Form(...), observaciones: str = Form(""), usuario: dict = Depends(obtener_usuario_actual)) -> dict:
    res = _caso_registrar_resultado().ejecutar(id_control, resultado, observaciones)
    if "error" in res:
        return res
    _auditar("control_calidad", id_control, "ACTUALIZAR", f"Resultado: {resultado}", usuario["sub"])
    return res


# ---------------------------------------------------------------------------
# Lotes / Trazabilidad
# ---------------------------------------------------------------------------
@app.get("/lotes", dependencies=[Depends(obtener_usuario_actual)])
def listar_lotes(page: int = 0, page_size: int = 0, q: str = "") -> dict:
    items, total = _repo_lotes().listar(page, page_size, q)
    if not page:
        page = 1
        page_size = total if total > 0 else 1
    return _paginar(items, total, page, page_size)


@app.get("/lotes/{codigo_lote}/trazabilidad", dependencies=[Depends(obtener_usuario_actual)])
def trazar_lote(codigo_lote: str) -> dict:
    res = _caso_trazar_lote().ejecutar(codigo_lote)
    if not res:
        return {"error": f"Lote '{codigo_lote}' no encontrado"}
    return res


@app.post("/ordenes/{id_orden}/lotes", dependencies=[Depends(requerir_rol("admin", "produccion"))])
def crear_lote(id_orden: str, codigo_lote: str = Form(...), cantidad_producida: float = Form(...), usuario: dict = Depends(obtener_usuario_actual)) -> dict:
    res = _caso_crear_lote().ejecutar(id_orden, codigo_lote, cantidad_producida)
    if "error" in res:
        return res
    _auditar("lote", res["id"], "CREAR", f"Lote '{codigo_lote}' para orden {id_orden[:8]}…", usuario["sub"])
    return res


@app.get("/ordenes/{id_orden}/lotes", dependencies=[Depends(obtener_usuario_actual)])
def listar_lotes_orden(id_orden: str) -> list[dict]:
    return _repo_lotes().obtener_por_orden(id_orden)


@app.put("/lotes/{id_lote}/estado", dependencies=[Depends(requerir_rol("admin", "produccion"))])
def actualizar_estado_lote(id_lote: str, estado: str = Form(...), usuario: dict = Depends(obtener_usuario_actual)) -> dict:
    if estado not in ("activo", "bloqueado", "liberado"):
        return {"error": "Estado debe ser 'activo', 'bloqueado' o 'liberado'"}
    if not _repo_lotes().actualizar_estado(id_lote, estado):
        return {"error": f"Lote '{id_lote}' no encontrado"}
    _auditar("lote", id_lote, "ACTUALIZAR", f"Estado: {estado}", usuario["sub"])
    return {"mensaje": f"Lote actualizado a '{estado}'"}


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    uvicorn.run("src.main:app", reload=True)
